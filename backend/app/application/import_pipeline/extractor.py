"""
File extraction layer for the import pipeline.
Returns plain dicts — no SQLAlchemy/DB imports here.
"""

from __future__ import annotations
import hashlib
from pathlib import Path
from typing import Any


# ── Checksum ────────────────────────────────────────────────────────────────

def compute_checksum(file_bytes: bytes) -> str:
    """Return SHA-256 hex digest of file bytes."""
    return hashlib.sha256(file_bytes).hexdigest()


# ── Excel extraction ─────────────────────────────────────────────────────────

def extract_excel(
    file_bytes: bytes,
    source_file_id: str,
    max_sheets: int = 20,
    max_rows: int = 2000,
) -> list[dict[str, Any]]:
    """
    Extract cells from an Excel workbook using openpyxl.
    Skips empty cells and lock/temp files.
    Returns a list of raw extraction dicts.
    """
    import openpyxl  # lazy import — not available in test env without install
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    extractions: list[dict] = []

    for sheet_name in wb.sheetnames[:max_sheets]:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=False), start=1):
            for cell in row:
                if cell.value is None:
                    continue
                raw_value = str(cell.value).strip()
                if not raw_value:
                    continue
                extractions.append({
                    "source_file_id": source_file_id,
                    "sheet_name": sheet_name,
                    "page_number": None,
                    "row_number": cell.row,
                    "column_name": cell.column_letter,
                    "cell_reference": cell.coordinate,
                    "raw_value": raw_value,
                    "extraction_method": "xlsx_cell",
                })

    wb.close()
    return extractions


# ── PDF extraction ───────────────────────────────────────────────────────────

def extract_pdf(
    file_bytes: bytes,
    source_file_id: str,
) -> list[dict[str, Any]]:
    """
    Extract text blocks from a PDF using pdfplumber.
    Each text line on each page becomes one raw extraction dict.
    Falls back to returning a single 'extraction_failed' record on error.
    """
    try:
        import pdfplumber  # lazy import
        from io import BytesIO

        extractions: list[dict] = []

        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                for line_num, line in enumerate(text.splitlines(), start=1):
                    raw_value = line.strip()
                    if not raw_value:
                        continue
                    extractions.append({
                        "source_file_id": source_file_id,
                        "sheet_name": None,
                        "page_number": page_num,
                        "row_number": line_num,
                        "column_name": None,
                        "cell_reference": None,
                        "raw_value": raw_value,
                        "extraction_method": "pdf_text",
                    })

        return extractions

    except Exception as exc:
        return [{
            "source_file_id": source_file_id,
            "sheet_name": None,
            "page_number": 1,
            "row_number": 1,
            "column_name": None,
            "cell_reference": None,
            "raw_value": f"[extraction_failed: {exc}]",
            "extraction_method": "pdf_text",
        }]


# ── Router ───────────────────────────────────────────────────────────────────

def extract_file(
    file_bytes: bytes,
    file_name: str,
    source_file_id: str,
) -> tuple[list[dict], str]:
    """
    Route to the correct extractor based on file extension.
    Returns (extractions, detected_file_type).
    """
    lower = file_name.lower()
    if lower.endswith((".xlsx", ".xlsb", ".xls", ".xlsm", ".csv")):
        return extract_excel(file_bytes, source_file_id), "xlsx"
    elif lower.endswith(".pdf"):
        return extract_pdf(file_bytes, source_file_id), "pdf"
    else:
        return [], "unknown"


def is_lock_file(file_name: str) -> bool:
    """Return True for Office temporary lock files (~$*.xlsx etc.)."""
    return Path(file_name).name.startswith("~$")
