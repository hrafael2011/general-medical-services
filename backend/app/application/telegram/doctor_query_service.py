"""Deterministic doctor queries for filtered Telegram questions."""

import io
import logging
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from backend.app.application.telegram.sanitize import display_value, format_rows
from backend.app.application.telegram.types import AgentResult
from backend.app.infrastructure.db.models.catalogs import DepartmentModel, RankModel
from backend.app.infrastructure.db.models.doctors import DoctorModel

logger = logging.getLogger(__name__)

_COLUMN_TITLES = {
    "name": "Nombre",
    "sex": "Sexo",
    "rank": "Rango",
    "total": "Total",
}

_SEX_LABELS = {
    "male": "Masculino",
    "female": "Femenino",
}


def _sorted_filters(filters: dict[str, Any]) -> dict[str, Any]:
    return {key: filters[key] for key in sorted(filters)}


def _possible_duplicate_names(rows: list[dict]) -> list[dict[str, Any]]:
    counts: dict[str, int] = {}
    for row in rows:
        name = str(row.get("name", "")).strip()
        if name:
            counts[name] = counts.get(name, 0) + 1
    return [
        {"name": name, "count": count}
        for name, count in sorted(counts.items())
        if count > 1
    ]


class DoctorQueryService:
    """Runs controlled doctor queries from resolved entities."""

    def __init__(self, session: Session) -> None:
        self._session = session

    def execute(self, user_text: str, resolved: dict[str, Any]) -> AgentResult | None:
        """Return an AgentResult when the request is a supported doctor query."""
        filters = self._filters_from_resolved(resolved)
        if not filters:
            return None

        is_export = self._is_export_request(user_text)
        operation = "list" if is_export else self._operation_from_text(user_text, filters)
        logger.info(
            "Doctor query deterministic route selected",
            extra={
                "telegram_event": "doctor_query_route",
                "match_type": "deterministic",
                "operation": operation,
                "requested_filters": _sorted_filters(filters),
                "is_export": is_export,
            },
        )
        if operation == "count_by_sex":
            rows, columns = self._count_by_sex(filters)
        elif operation == "count":
            rows, columns = self._count(filters)
        else:
            rows, columns = self._list(filters)
        possible_duplicates = _possible_duplicate_names(rows)

        validation = self._validate_result_filters(rows, filters, operation)
        if not validation["ok"]:
            logger.warning(
                "Doctor query filter validation failed",
                extra={
                    "telegram_event": "doctor_query_validation_failed",
                    "operation": operation,
                    "requested_filters": _sorted_filters(filters),
                    "validation_error": validation.get("error"),
                },
            )
            return AgentResult(
                response_text="No pude validar que todos los filtros pedidos fueron aplicados.",
                agent_action="validation_error",
                tool_name="doctor_query_service",
                tool_entities={
                    "requested_filters": _sorted_filters(filters),
                    "operation": operation,
                },
                tool_result=validation,
            )

        if is_export:
            return self._export_result(
                user_text,
                filters,
                rows,
                columns,
                validation,
                possible_duplicates,
            )

        logger.info(
            "Doctor query completed",
            extra={
                "telegram_event": "doctor_query_completed",
                "match_type": "deterministic",
                "operation": operation,
                "applied_filters": _sorted_filters(filters),
                "row_count": len(rows),
                "possible_duplicate_name_count": len(possible_duplicates),
            },
        )
        response_text = format_rows(rows, columns)
        if possible_duplicates:
            dup_lines = "\n".join(
                f"  - {d['name']} ({d['count']} registros)"
                for d in possible_duplicates[:10]
            )
            response_text += (
                f"\n\nPosibles duplicados por nombre ({len(possible_duplicates)}):\n{dup_lines}"
            )
        return AgentResult(
            response_text=response_text,
            agent_action="query",
            tool_name="doctor_query_service",
            tool_entities={
                "requested_filters": _sorted_filters(filters),
                "applied_filters": _sorted_filters(filters),
                "operation": operation,
            },
            tool_result={
                "ok": True,
                "source": "deterministic_doctor_query",
                "row_count": len(rows),
                "validated_filters": validation["validated_filters"],
                "possible_duplicate_names": possible_duplicates,
                "data": {"columns": columns, "rows": rows},
            },
        )

    def _filters_from_resolved(self, resolved: dict[str, Any]) -> dict[str, Any]:
        filters: dict[str, Any] = {}

        rank = resolved.get("rank")
        if isinstance(rank, dict) and rank.get("normalized_name"):
            filters["rank"] = rank["normalized_name"]

        department = resolved.get("department")
        if isinstance(department, dict) and department.get("normalized_name"):
            filters["department"] = department["normalized_name"]

        sex = resolved.get("sex")
        if isinstance(sex, list):
            filters["sex"] = sex
        elif sex:
            filters["sex"] = [sex]

        return filters

    def _operation_from_text(self, user_text: str, filters: dict[str, Any]) -> str:
        text = user_text.lower()
        asks_count = any(word in text for word in ("cuanto", "cuantos", "cuanta", "cuantas"))
        sex_values = filters.get("sex") or []
        if asks_count and len(sex_values) > 1:
            return "count_by_sex"
        if asks_count:
            return "count"
        return "list"

    def _is_export_request(self, user_text: str) -> bool:
        text = user_text.lower()
        return any(
            word in text
            for word in (
                "exporta",
                "exportar",
                "reporte",
                "pdf",
                "excel",
                "xlsx",
                "informacion",
                "información",
            )
        )

    def _format_from_text(self, user_text: str) -> str:
        text = user_text.lower()
        if "excel" in text or "xlsx" in text:
            return "excel"
        return "pdf"

    def _title_for_filters(self, filters: dict[str, Any]) -> str:
        parts = ["MEDICOS FILTRADOS"]
        if filters.get("rank"):
            parts.append(str(filters["rank"]).upper())
        sex_values = filters.get("sex") or []
        if len(sex_values) == 1:
            parts.append(_SEX_LABELS.get(sex_values[0], sex_values[0]).upper())
        elif len(sex_values) > 1:
            parts.append("POR SEXO")
        if filters.get("department"):
            parts.append(str(filters["department"]).upper())
        return " - ".join(parts)

    def _export_result(
        self,
        user_text: str,
        filters: dict[str, Any],
        rows: list[dict],
        columns: list[str],
        validation: dict[str, Any],
        possible_duplicates: list[dict[str, Any]],
    ) -> AgentResult:
        if not rows:
            return AgentResult(
                response_text="No se encontraron resultados para generar el reporte.",
                agent_action="export",
                tool_name="doctor_query_service",
                tool_entities={
                    "requested_filters": _sorted_filters(filters),
                    "applied_filters": _sorted_filters(filters),
                    "operation": "export",
                },
                tool_result={
                    "ok": True,
                    "source": "deterministic_doctor_query",
                    "row_count": len(rows),
                    "validated_filters": validation["validated_filters"],
                    "possible_duplicate_names": possible_duplicates,
                    "data": {"columns": columns, "rows": rows},
                },
            )

        fmt = self._format_from_text(user_text)
        title = self._title_for_filters(filters)
        if fmt == "excel":
            document_bytes = self._build_excel(rows, columns)
            filename = "MEDICOS_FILTRADOS.xlsx"
        else:
            document_bytes = self._build_pdf(rows, columns, title)
            filename = "MEDICOS_FILTRADOS.pdf"

        logger.info(
            "Doctor query export completed",
            extra={
                "telegram_event": "doctor_query_export_completed",
                "match_type": "deterministic",
                "operation": "export",
                "applied_filters": _sorted_filters(filters),
                "export_format": fmt,
                "row_count": len(rows),
                "possible_duplicate_name_count": len(possible_duplicates),
                "document_filename": filename,
            },
        )
        return AgentResult(
            response_text=(
                f"Aquí tienes el reporte solicitado. "
                f"({len(rows)} registros, {fmt.upper()})."
            ),
            document_bytes=document_bytes,
            document_filename=filename,
            agent_action="export",
            tool_name="doctor_query_service",
            tool_entities={
                "requested_filters": _sorted_filters(filters),
                "applied_filters": _sorted_filters(filters),
                "operation": "export",
                "export_format": fmt,
            },
            tool_result={
                "ok": True,
                "source": "deterministic_doctor_query",
                "row_count": len(rows),
                "validated_filters": validation["validated_filters"],
                "possible_duplicate_names": possible_duplicates,
                "data": {"columns": columns, "rows": rows},
            },
        )

    def _validate_result_filters(
        self,
        rows: list[dict],
        filters: dict[str, Any],
        operation: str,
    ) -> dict[str, Any]:
        validated = [key for key in ("rank", "sex", "department") if key in filters]
        if not rows or operation in ("count", "count_by_sex"):
            return {"ok": True, "validated_filters": validated}

        sex_values = filters.get("sex") or []
        expected_rank = str(filters.get("rank", "")).lower()

        for row in rows:
            if sex_values and set(sex_values) != {"male", "female"}:
                if row.get("sex") not in sex_values:
                    return {
                        "ok": False,
                        "error": "sex_filter_not_applied",
                        "row": row,
                        "validated_filters": validated,
                    }
            if expected_rank and str(row.get("rank", "")).lower() != expected_rank:
                return {
                    "ok": False,
                    "error": "rank_filter_not_applied",
                    "row": row,
                    "validated_filters": validated,
                }
        return {"ok": True, "validated_filters": validated}

    def _build_pdf(self, rows: list[dict], columns: list[str], title: str) -> bytes:
        from reportlab.lib.units import cm

        from backend.app.application.reports.pdf_templates import generate_doctor_list_pdf

        header_titles = [
            _COLUMN_TITLES.get(column, column.replace("_", " ").title())
            for column in columns
        ]
        pdf_rows = [
            {
                title: display_value(column, row.get(column, ""))
                for title, column in zip(header_titles, columns, strict=True)
            }
            for row in rows
        ]
        col_widths = [max(2.5 * cm, len(title) * 0.18 * cm) for title in header_titles]
        col_widths = [min(width, 6 * cm) for width in col_widths]
        return generate_doctor_list_pdf(
            pdf_rows,
            title=title,
            columns=header_titles,
            col_widths=col_widths,
        )

    def _build_excel(self, rows: list[dict], columns: list[str]) -> bytes:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        ws.append([
            _COLUMN_TITLES.get(column, column.replace("_", " ").title())
            for column in columns
        ])
        for row in rows:
            ws.append([display_value(column, row.get(column, "")) for column in columns])
        from openpyxl.utils import get_column_letter
        for i, column in enumerate(columns, start=1):
            letter = get_column_letter(i)
            title = _COLUMN_TITLES.get(column, column.replace("_", " ").title())
            ws.column_dimensions[letter].width = max(12, len(title) + 4)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _base_conditions(self, filters: dict[str, Any]) -> list:
        conditions = [
            DoctorModel.active.is_(True),
            DoctorModel.service_active.is_(True),
        ]
        if filters.get("rank"):
            conditions.append(func.lower(RankModel.normalized_name) == str(filters["rank"]).lower())
        if filters.get("department"):
            conditions.append(
                func.lower(DepartmentModel.normalized_name) == str(filters["department"]).lower()
            )
        sex_values = filters.get("sex") or []
        if sex_values and set(sex_values) != {"male", "female"}:
            conditions.append(DoctorModel.sex.in_(sex_values))
        return conditions

    def _join_catalogs(self, stmt, filters: dict[str, Any]):
        stmt = stmt.outerjoin(RankModel, DoctorModel.rank_id == RankModel.id)
        if filters.get("department"):
            stmt = stmt.join(DepartmentModel, DoctorModel.department_id == DepartmentModel.id)
        return stmt

    def _count(self, filters: dict[str, Any]) -> tuple[list[dict], list[str]]:
        stmt = select(
            func.count(func.distinct(DoctorModel.id)).label("total")
        ).select_from(DoctorModel)
        stmt = self._join_catalogs(stmt, filters).where(*self._base_conditions(filters))
        row = self._session.execute(stmt).mappings().one()
        return [dict(row)], ["total"]

    def _count_by_sex(self, filters: dict[str, Any]) -> tuple[list[dict], list[str]]:
        stmt = (
            select(
                DoctorModel.sex.label("sex"),
                func.count(func.distinct(DoctorModel.id)).label("total"),
            )
            .select_from(DoctorModel)
            .group_by(DoctorModel.sex)
            .order_by(DoctorModel.sex)
        )
        stmt = self._join_catalogs(stmt, filters).where(*self._base_conditions(filters))
        rows = [dict(row) for row in self._session.execute(stmt).mappings().all()]
        return rows, ["sex", "total"]

    def _list(self, filters: dict[str, Any]) -> tuple[list[dict], list[str]]:
        columns = ["name", "sex", "rank"]
        stmt = (
            select(
                DoctorModel.id.label("_doctor_id"),
                DoctorModel.name.label("name"),
                DoctorModel.sex.label("sex"),
                RankModel.name.label("rank"),
            )
            .select_from(DoctorModel)
            .order_by(DoctorModel.name)
        )
        stmt = self._join_catalogs(stmt, filters).where(*self._base_conditions(filters))
        seen_ids = set()
        rows = []
        for row in self._session.execute(stmt).mappings().all():
            row_dict = dict(row)
            doctor_id = row_dict.pop("_doctor_id")
            if doctor_id in seen_ids:
                continue
            seen_ids.add(doctor_id)
            rows.append(row_dict)
        return rows, columns
